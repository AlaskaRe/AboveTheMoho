from openpyxl import load_workbook

Point = {}
stratums_lib = []

wb = load_workbook(filename="./RawData/RawPointData.xlsx", data_only=True)
wksht_pt = wb["POSTION"]
wksht_lib = wb["LAYER_LIBARY"]
wksht_args = wb["POINT_ARGS"]

# read the worksheet"LAYER_LIBARY", get the information of the layers
for row in wksht_lib.iter_rows(min_row=2, max_row=20, max_col=3):
    if row[0].value != None:

        T = (row[0].value, row[1].value, row[2].value)
        stratums_lib.append(T)
# If I define the layerInformation outside the for loop, then every point get the adress of this list,
# then everytime I give the value to the point , it will modify the layerinformation, every point has the same list value
# layerInformation = [None for i in range(len(stratums_lib))]

# get the point data from the sheet
for r in wksht_pt.iter_rows(min_row=2, max_col=3):
    if r[0].value == None:
        break
    layerInformation = [None for i in range(len(stratums_lib))]
    Point[r[0].value, r[1].value, r[2].value] = layerInformation

for r in wksht_args.iter_rows(min_row=2, max_col=13):

    if r[0].value == None:
        break
    else:
        # Read the row and get the value of every stratum for each point.
        kywd = (r[2].value, r[3].value, r[12].value)
        da = [r[7].value, r[8].value, r[9].value]
        idx = None
        for i in stratums_lib:
            if i[0] == r[11].value:
                idx = stratums_lib.index(i)
                break

        # give the stratum value to the dict if not exists, else add another value to this list
        sttms = Point[kywd][idx]

        if sttms == None:
            Point[kywd][idx] = da

        elif sttms != None:
            t = []
            t.append(sttms)
            t.append(da)
            Point[kywd][idx] = t


print("hello ")
